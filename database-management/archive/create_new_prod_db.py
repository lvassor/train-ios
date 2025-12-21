#!/usr/bin/env python3
"""
Create new production database with:
1. Trimmed instructions (single sentence per step)
2. Removed duplicate EX030/EX031
3. Resequenced exercise IDs
4. Video mapping table
5. Updated is_in_programme flags
6. Fixed lookup sheets (remove header rows)
"""

import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# CONFIGURATION - Exercises to set is_in_programme = 0 (per update.txt)
# ============================================================================
EXERCISES_TO_DISABLE = {
    'Single Leg Kettlebell Calf Raise',  # False
    'Glute Kickback',                     # False
    'Kettlebell Walking Lunge',           # False
    'Kettlebell Bent Over Row',           # False
    'Kettlbell Bulgarian Split Squat',    # False (typo in DB)
    'Kettlebell Bulgarian Split Squat',   # False (correct spelling if exists)
    'Elevated Kettlbell Bulgarian Split Squat',  # False (typo in DB)
    'Elevated Kettlebell Bulgarian Split Squat', # False (correct spelling if exists)
    'Kettlebell Goblet Squat (heels elevated)',  # False
    'Pallof Press',                       # False - no video available
}

# Exercises that use variant videos - note what variant is used
VARIANT_VIDEO_NOTES = {
    'Dumbbell Back Extension': 'Uses bodyweight back extension video',
    'Kettlebell Back Extension': 'Uses bodyweight back extension video',
    'Seated Calf Raise': 'Uses plate-loaded variation',  # for pin-loaded version
    'Standing Calf Raise': 'Uses pin-loaded variation',  # for plate-loaded version
    'Calf Press': 'Uses plate-loaded variation',  # for pin-loaded version
    'Ab Wheel Rollout': 'Uses kneeling variation',
    'Farmer\'s Walk': 'Uses dumbbell variation',
    'Kettlebell Farmer\'s Walk': 'Uses dumbbell Farmer\'s Walk video',
    'Single Leg Glute Bridge Hold': 'Uses dynamic single leg bridge video',
    'Elevated Dumbbell Reverse Lunge': 'Uses standard reverse lunge video',
    'Elevated Kettlebell Reverse Lunge': 'Uses kettlebell reverse lunge video',
    'Copenhagen Plank (bent knee)': 'Uses Side Plank Hip Adduction (bent knee)',
    'Copenhagen Plank (straight leg)': 'Uses Side Plank Hip Adduction',
}


def trim_instruction_step(step_text):
    """Trim a step to its first sentence only."""
    match = re.match(r'^(Step \d+:)\s*(.+)$', step_text.strip())
    if not match:
        return step_text.strip()

    prefix = match.group(1)
    content = match.group(2)

    # Get first sentence - split on period followed by space and capital letter
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', content)
    first_sentence = sentences[0].strip()

    # Ensure it ends with a period
    if not first_sentence.endswith(('.', '!', '?')):
        first_sentence += '.'

    return f"{prefix} {first_sentence}"


def trim_instructions(full_instructions):
    """Trim all steps in the instructions to single sentences."""
    if pd.isna(full_instructions) or not full_instructions:
        return ""

    steps = full_instructions.split('\n')
    trimmed_steps = [trim_instruction_step(step) for step in steps if step.strip()]
    return '\n'.join(trimmed_steps)


def resequence_ids(df, start_id=1):
    """Resequence exercise IDs after removing duplicates."""
    new_ids = [f"EX{str(i).zfill(3)}" for i in range(start_id, start_id + len(df))]
    df['exercise_id'] = new_ids
    return df


def main():
    print("=" * 60)
    print("Creating new production database")
    print("=" * 60)

    # Load all sheets from original Excel
    excel_path = 'archive/train_exercise_database_prod.xlsx'
    xl = pd.ExcelFile(excel_path)

    # Read master sheet
    df = pd.read_excel(xl, sheet_name='master')
    print(f"\nOriginal exercise count: {len(df)}")

    # Show duplicates before removal
    print(f"\nRemoving duplicates EX030 ({df.loc[df['exercise_id'] == 'EX030', 'display_name'].values[0]}) and EX031 ({df.loc[df['exercise_id'] == 'EX031', 'display_name'].values[0]})")

    # Remove EX030 and EX031 (duplicates of decline presses)
    df = df[~df['exercise_id'].isin(['EX030', 'EX031'])]
    print(f"After removing duplicates: {len(df)}")

    # Resequence IDs
    df = resequence_ids(df)
    print(f"Resequenced IDs from EX001 to EX{str(len(df)).zfill(3)}")

    # Trim instructions
    df['instructions'] = df['instructions'].apply(trim_instructions)
    print("Trimmed instructions to single sentences per step")

    # Update is_in_programme for exercises to disable
    exercises_disabled = []
    for idx, row in df.iterrows():
        display_name = row['display_name']
        if display_name in EXERCISES_TO_DISABLE:
            df.at[idx, 'is_in_programme'] = 0
            exercises_disabled.append(f"{row['exercise_id']}: {display_name}")

    print(f"\nSet is_in_programme=0 for {len(exercises_disabled)} exercises:")
    for ex in exercises_disabled:
        print(f"  - {ex}")

    # Load original workbook to preserve structure
    output_path = 'train_exercise_database_v2.xlsx'
    wb = load_workbook(excel_path)

    # Update master sheet with new data
    ws = wb['master']

    # Clear existing data (keep headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write new data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # ========================================================================
    # Fix lookup sheets - remove header rows (row 1 should be data, not headers)
    # ========================================================================
    print("\nFixing lookup sheets...")
    lookup_sheets = ['canonical_lookup', 'muscle_lookup', 'equipment_category_lookup', 'equipment_specific_lookup']

    for sheet_name in lookup_sheets:
        if sheet_name in wb.sheetnames:
            ws_lookup = wb[sheet_name]
            # Check if row 1 looks like a header (contains text like 'canonical_name', 'muscle', etc.)
            first_cell = ws_lookup.cell(row=1, column=1).value
            if first_cell and isinstance(first_cell, str) and ('_' in first_cell or first_cell.lower() == first_cell):
                # This looks like a header row - delete it
                ws_lookup.delete_rows(1)
                print(f"  - Removed header row from {sheet_name}")
            else:
                print(f"  - {sheet_name} already has data in row 1")

    # Save workbook
    wb.save(output_path)
    print(f"\nSaved new database to: {output_path}")

    # ========================================================================
    # Create video mapping table
    # ========================================================================
    print("\n" + "=" * 60)
    print("Creating video mapping table")
    print("=" * 60)

    # Read the actual purchase order (ground truth IDs)
    purchase_df = pd.read_csv('archive/luke_vassor_purchase_order.csv')
    print(f"Loaded {len(purchase_df)} items from purchase order")

    # Create a lookup from purchase order Name to ID
    purchase_lookup = {}
    for _, row in purchase_df.iterrows():
        purchase_lookup[row['Name']] = {'id': str(row['ID']), 'type': row['type']}

    # Manual mappings for exercises that don't fuzzy match well
    manual_mappings = {
        'Back Iso Hold': {'id': '8895', 'type': 'img'},
        'Dips': {'id': '143012', 'type': 'vid'},
        'Copenhagen Plank (straight leg)': {'id': '177512', 'type': 'vid'},
        'Copenhagen Plank (bent knee)': {'id': '346012', 'type': 'vid'},
        'Body Weight Glute Bridge': {'id': '801512', 'type': 'vid'},
        'T-Bar Row': {'id': '320012', 'type': 'vid'},
        'Single Leg Dumbbell Calf Raise': {'id': '572512', 'type': 'vid'},
        'Standing Dumbbell Calf Raise': {'id': '041712', 'type': 'vid'},
        'Dumbbell Back Extension': {'id': '547112', 'type': 'vid'},
        'Kettlebell Back Extension': {'id': '547112', 'type': 'vid'},
        'Kettlebell Farmer\'s Walk': {'id': '213312', 'type': 'vid'},
        'Single Leg Glute Bridge Hold': {'id': '6001', 'type': 'img'},  # Screenshot of hold
        'Single Leg Glute Bridge': {'id': '600012', 'type': 'vid'},
        'Elevated Dumbbell Reverse Lunge': {'id': '732612', 'type': 'vid'},
        'Elevated Kettlebell Reverse Lunge': {'id': '357512', 'type': 'vid'},
        # Bench press variations - need exact IDs from purchase order
        'Barbell Bench Press': {'id': '002512', 'type': 'vid'},
        'Dumbbell Bench Press': {'id': '028912', 'type': 'vid'},
        'Dumbbell Floor Press': {'id': '742812', 'type': 'vid'},
        'Cable Chest Press': {'id': '335912', 'type': 'vid'},
        'Incline Barbell Bench Press': {'id': '004712', 'type': 'vid'},
        'Incline Dumbbell Bench Press': {'id': '031412', 'type': 'vid'},
        'Decline Barbell Bench Press': {'id': '003312', 'type': 'vid'},
        'Decline Dumbbell Bench Press': {'id': '030112', 'type': 'vid'},
        'Close Grip Bench Press': {'id': '003012', 'type': 'vid'},
        # Other exercises that need explicit mapping
        'Plank': {'id': '5159', 'type': 'img'},
        'Side Plank': {'id': '0715', 'type': 'img'},
        'Ab Wheel Rollout': {'id': '777112', 'type': 'vid'},
        'Farmer\'s Walk': {'id': '213312', 'type': 'vid'},
        # Push-up variations
        'Push Up': {'id': '066212', 'type': 'vid'},
        'Incline Push Up': {'id': '906412', 'type': 'vid'},
        'Decline Push Up': {'id': '027912', 'type': 'vid'},
        'Pike Push Up': {'id': '742612', 'type': 'vid'},
        'Diamond Push Up': {'id': '269712', 'type': 'vid'},
        # Core exercises
        'Cable Crunch': {'id': '236912', 'type': 'vid'},
        'Pallof Press': {'id': '', 'type': ''},  # No video available
        # Deadlift variations
        'Barbell Landmine Romanian Deadlift': {'id': '424712', 'type': 'vid'},
        # Lunge variations
        'Dumbbell Reverse Lunge': {'id': '038112', 'type': 'vid'},
        # Row variations
        'Dumbbell Single Arm Row': {'id': '239312', 'type': 'vid'},
        'Chest Supported Row': {'id': '332012', 'type': 'vid'},
        # Shoulder press variations
        'Barbell Overhead Press': {'id': '532812', 'type': 'vid'},
        'Dumbbell Overhead Press': {'id': '040512', 'type': 'vid'},
        # Tricep exercises
        'Dumbbell Skull Crusher': {'id': '1043212', 'type': 'vid'},
        'Overhead Dumbbell Extension': {'id': '615812', 'type': 'vid'},
    }

    # Get list of files in demo_videos
    demo_videos_path = '../demo_videos'
    video_files = [f for f in os.listdir(demo_videos_path) if f.endswith(('.mp4', '.png'))]
    print(f"Found {len(video_files)} media files in demo_videos")

    # Create mapping table
    mapping_data = []

    # Build a lookup from file ID to filename
    file_lookup = {}
    for f in video_files:
        file_id = f.split('-')[0] if '-' in f else f.split('.')[0]
        file_lookup[file_id] = f
        file_lookup[file_id.lstrip('0')] = f

    print(f"Built file lookup with {len(file_lookup)} entries")

    # Process each exercise from the new database
    for _, row in df.iterrows():
        new_id = row['exercise_id']
        display_name = row['display_name']

        supplier_id = ''
        file_type = ''
        filename = ''
        note = ''

        # Check manual mappings first
        if display_name in manual_mappings:
            supplier_id = manual_mappings[display_name]['id']
            file_type = manual_mappings[display_name]['type']
        else:
            # Try exact match first
            for po_name, po_info in purchase_lookup.items():
                if display_name.lower() in po_name.lower() or po_name.lower() in display_name.lower():
                    supplier_id = po_info['id']
                    file_type = po_info['type']
                    break

            # If no match, try fuzzy matching
            if not supplier_id:
                from difflib import SequenceMatcher
                best_score = 0
                best_match = None
                for po_name, po_info in purchase_lookup.items():
                    score = SequenceMatcher(None, display_name.lower(), po_name.lower()).ratio()
                    if score > best_score and score > 0.5:
                        best_score = score
                        best_match = (po_name, po_info)
                if best_match:
                    supplier_id = best_match[1]['id']
                    file_type = best_match[1]['type']

        # Find matching filename using supplier ID
        if supplier_id:
            supplier_id_str = str(supplier_id)

            if file_type == 'vid':
                if len(supplier_id_str) >= 3:
                    base_id = supplier_id_str[:-2]
                    target_id = base_id + '1201'
                else:
                    target_id = supplier_id_str + '1201'
            elif file_type == 'img':
                target_id = supplier_id_str + '1101'
            else:
                if len(supplier_id_str) >= 3:
                    base_id = supplier_id_str[:-2]
                    target_id = base_id + '1201'
                else:
                    target_id = supplier_id_str + '1201'

            target_id_padded = target_id.zfill(8)

            if target_id_padded in file_lookup:
                filename = file_lookup[target_id_padded]
            elif target_id in file_lookup:
                filename = file_lookup[target_id]
            elif target_id.lstrip('0') in file_lookup:
                filename = file_lookup[target_id.lstrip('0')]
            elif target_id_padded.lstrip('0') in file_lookup:
                filename = file_lookup[target_id_padded.lstrip('0')]

        # Determine note based on status
        if display_name in EXERCISES_TO_DISABLE:
            note = 'Unavailable'
        elif display_name in VARIANT_VIDEO_NOTES:
            note = VARIANT_VIDEO_NOTES[display_name]
        elif not filename and not supplier_id:
            note = 'Unavailable'

        mapping_data.append({
            'exercise_id': new_id,
            'display_name': display_name,
            'supplier_id': supplier_id,
            'media_type': file_type,
            'filename': filename,
            'note': note
        })

    # Create mapping DataFrame
    mapping_df = pd.DataFrame(mapping_data)

    # Save mapping table (no exact_match column)
    mapping_path = 'exercise_video_mapping.csv'
    mapping_df.to_csv(mapping_path, index=False)
    print(f"\nSaved video mapping to: {mapping_path}")

    # Summary
    with_files = len(mapping_df[mapping_df['filename'] != ''])
    without_files = len(mapping_df[mapping_df['filename'] == ''])
    unavailable = len(mapping_df[mapping_df['note'] == 'Unavailable'])
    variants = len(mapping_df[mapping_df['note'].str.contains('Uses', na=False)])

    print(f"\nMapping summary:")
    print(f"  - Exercises with media files: {with_files}")
    print(f"  - Exercises without media files: {without_files}")
    print(f"  - Marked as Unavailable: {unavailable}")
    print(f"  - Using variant videos: {variants}")

    # Show unavailable exercises
    print("\nUnavailable exercises (is_in_programme=0):")
    for _, row in mapping_df[mapping_df['note'] == 'Unavailable'].iterrows():
        print(f"  - {row['exercise_id']}: {row['display_name']}")

    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == '__main__':
    main()
